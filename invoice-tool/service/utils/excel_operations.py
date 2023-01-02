import os
import openpyxl
import itertools
import service
import logging

from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError

logging.basicConfig(level = logging.DEBUG)

class ExcelException(Exception):
    """
        Exception class to handle excel(xlsx) file Exceptions
    """
    pass

def get_excel_headers()->list:
    """
    method to get headers of the invoice excel
    
    Returns
    -------
    headers of invoice excel
    """
    return ["From Delivery Date", "To Delivery Date", "Invoice Number", "Legal Entity Name", "Invoice Account", "Total Amount",
            "Currency", "Credit Bank Account", "Debit Bank Account", "Invoice Date",
            "Due Date", "Markets", "Reference Number", "Country/Region", "Paid"]

def create_excel_header(ws):
    """
    adds header to the invoice excel

    Params
    -------
        ws:
            worksheet of the invoice excel
    """
    first_row=1
    excel_headers= get_excel_headers() 
    for header in excel_headers:
        col= excel_headers.index(header) 
        current_cell = ws.cell(first_row,col+1)
        try:
            current_cell.value= header
        except IllegalCharacterError as e:
            raise ExcelException(f"error occured while creating excel:{e}")
        current_cell.font = Font(bold=True)
        current_cell.alignment = Alignment(wrap_text=True, vertical='center')

def format_excel_cells(ws):
    """
    adding dimension styles to all cells in invoice excel worksheet

    Params
    -------
        ws:
            worksheet of the invoice excel
    """
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=11)
    ws.column_dimensions = dim_holder
    

def excel_merge_execution(src_name:str, dest_ws):
    """
    takes two invoice files and performs merging

    Params
    -------
        src_name:str
            name of the invoice file that should be merged into another destination invoice
        dest_ws:
            worksheet of the invoice file and all other invoices should be merged into this
    """
    source_wb = load_workbook(src_name)
    source_sheet = source_wb.active
    max_row = dest_ws.max_row
    for row in source_sheet.rows:
            for cell in row:
                if cell.row !=1:
                    dest_ws.cell(cell.row+max_row-1,cell.column).value = cell.value

def merge_excel(dest_file_name:str):
    """
    initiates invoice file merging by creating yearly invoice and copying first invoice file

    Params
    -------
        dest_file_name: str
            the name of final yearly invoice excel
    """
    invoice_path = service.INVOICE_YEARLY_ROOT_PATH
    dest_folder = service.INVOICE_MERGE_ROOT_PATH
    if not os.path.exists(dest_folder):
        os.mkdir(dest_folder)
    dest_path = f"{dest_folder}/{dest_file_name}"
    files = [file for file in os.listdir(invoice_path)] 
    if len(files) > 0:
        try:
            files = sorted(files, key=lambda file: datetime.strptime(file.split("_")[1], "%b%y"))
        except IndexError:
            logging.exception("error occured while merging excels due to incorrect format of file")
            raise
        dest_wb= Workbook()
        dest_ws= dest_wb.active
        create_excel_header(dest_ws)
        for file in files:
            excel_merge_execution(f"{invoice_path}/{file}", dest_ws)
        format_excel_cells(dest_ws)
        dest_wb.save(dest_path)


def invoice_excel_generation(invoice_data:list, invoice_file_name:str): 
    """
    generates invoice excel file

    Params
    -------
        invoice_data: list
            data we get from nordpool external endpoint 
        invoice_file_name: str
            filename for invoice excel
    raises:
    -------
        ExcelException
    """
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.row_dimensions[1].height = 100
    create_excel_header(ws)
    ordered_list = ['Country' if item == 'Country/Region' else item.replace(" ", "") for item in get_excel_headers()]
    row=2
    for record in invoice_data:
        for _key,_value in record.items():
            if _key in ordered_list:
                if _key=="Markets" or _key=="SettlementTransactionIds":
                    _value = ' '.join(_value)
                if _key=="Paid":
                    if _value == True:
                        _value = "Yes"
                    else:
                        _value = "No"
                if _key in ["FromDeliveryDate", "ToDeliveryDate", "DueDate", "InvoiceDate"]:
                    date_value = "".join(itertools.takewhile(lambda x: x!="T", _value))
                    split_date_value = date_value.split('-')
                    _value = '/'.join(reversed(split_date_value))

                col=ordered_list.index(_key)
                try:
                    ws.cell(row,col+1).value = _value
                except IllegalCharacterError as e:
                    raise ExcelException(f"error occured while creating excel:{e}")
        row+=1
    
    format_excel_cells(ws)
    wb.save(filename = invoice_file_name)