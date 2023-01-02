import os
import json
import responses
import openpyxl
import pytest
from base.config import Config
from openpyxl.utils import get_column_letter
from marshmallow import ValidationError
from freezegun import freeze_time
from service.nordpool_script import main
from service.utils.nordpool_client import NordPoolApiClient, NordPoolException
from service.utils import excel_operations 
from service.utils.excel_operations import ExcelException, merge_excel, invoice_excel_generation
from unittest.mock import patch
from service import ROOT_PATH

NordPoolClient = NordPoolApiClient()
CONFIG = Config().get_config()

NORD_POOL_TOKEN_URL = CONFIG["NORDPOOL"]["TOKEN_URL"]
NORD_POOL_INVOICE_DOWNLOAD_ENDPOINT = CONFIG["NORDPOOL"]["INVOICE_DOWNLOAD_ENDPOINT"]

def load_json(file):
    filepath = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), file
    )
    with open(filepath, "r") as f:
        data = json.load(f)

    return data

invoice_data = load_json("assets/nordpool_invoice.json")

auth_response_mock = {
        "access_token": "8ShaEhIJNwsjfK82rt0jTaizlDzehALPhlmz9hUw5pGFv4m9XiPI0G",
        "token_type": "Bearer",
        "expires_in": 7200}

def read_excel(filename):
    #reading data from invoice excel
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    excel_invoice_data = []
    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))
    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
        excel_invoice_data.append(my_dict)
    wb.close()
    return excel_invoice_data

@pytest.mark.usefixtures("mock_set_client")    
@responses.activate
@freeze_time("2022-01-30")
def test_nordpool_script_success():
    responses.add(responses.POST, f"{NORD_POOL_TOKEN_URL}", json=auth_response_mock,
                  status=200)

    responses.add(responses.GET, f"{NORD_POOL_INVOICE_DOWNLOAD_ENDPOINT}?fromDeliveryDate=2022-01-01&toDeliveryDate=2022-01-29", json=load_json("assets/nordpool_invoice.json"),
                  status=200)
    # main creates invoice excel
    with patch('os.remove'):
        main()
    expected_invoice_data = load_json("assets/excel_invoice_data.json")
    assert read_excel("Invoice_Jan22_gesamt.xlsx") == expected_invoice_data
    os.remove("Invoice_Jan22_gesamt.xlsx")

@patch("service.INVOICE_YEARLY_ROOT_PATH", f"{ROOT_PATH}/tests/unit/invoice/yearly")
@patch("service.INVOICE_MERGE_ROOT_PATH", f"{ROOT_PATH}/tests/unit/invoice/merged")
def test_merge_excel():
    invoice_excel_name = f"{ROOT_PATH}/tests/unit/invoice/merged/Invoice_2022.xlsx"
    merge_excel("Invoice_2022.xlsx")
    expected_invoice_data = load_json("assets/excel_merge.json")
    merge_excel_data = read_excel(invoice_excel_name)
    assert merge_excel_data == expected_invoice_data
    os.remove(invoice_excel_name)

def test_merge_excel_index_error():
    with patch("os.listdir") as mock_files:
        mock_files.return_value = ["test.xlsx"]
        with pytest.raises(IndexError):
            merge_excel("Invoice_2022.xlsx")

@responses.activate
@freeze_time("2022-06-30")
def test_create_excel_header_failure(monkeypatch):
    invoice_content = load_json("assets/nordpool_invoice.json")
    def mock_headers(*args, **kwargs):
        return ["\x00"]
    monkeypatch.setattr(excel_operations, "get_excel_headers", mock_headers)
    with pytest.raises(ExcelException):
        invoice_excel_generation(invoice_content,"test_file")